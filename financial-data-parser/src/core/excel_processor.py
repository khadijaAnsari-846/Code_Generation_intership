import numpy as np
import pandas as pd
import os
import openpyxl
from IPython.display import display
from openpyxl import load_workbook,Workbook
import sqlite3
import re
from datetime import datetime
from decimal import Decimal
import locale
import re
import warnings
from collections import Counter
warnings.filterwarnings("ignore", category=UserWarning)  # Suppress the warning
from datetime import datetime, timedelta






# Step 2: File Reading Class

class ExcelProcessor:

  __work_book={}
  __sheet_data={}
  __sheet_name={}
  __sheet_data={}



  def __init__(self):
    self.__work_book={}
    self.__sheet_name={}
    self.__sheet_name={}
    self.__sheet_data={}


  def load_files(self, file_paths):
    for file_path in file_paths:
      try:
        self.__work_book[file_path]=load_workbook(file_path)
        print(f"File loaded completely:: {file_path}")
      except:
        print(f"Error loading file:{file_path}")
        continue


  def get_sheet_info(self):

    for work_book_name in self.__work_book.items():
      try:
        print(f"Main Work-Book: {work_book_name[0]}")
        self.__sheet_name[work_book_name[0]]=[]

        for sheet_name in work_book_name[1].sheetnames:
          print(f"Sub sheet:: {work_book_name[1][sheet_name]}")
          self.__sheet_name[work_book_name[0]].append(work_book_name[1][sheet_name].title)
          print(f"Dimentions of sheet is::{(work_book_name[1][sheet_name].max_row)} x {(work_book_name[1][sheet_name].max_column)}")
          print(f"All column name::")
          column_names = [
            cell.value for rows in work_book_name[1][sheet_name].iter_rows(min_row=1, max_row=1)
            for cell in rows if cell.value is not None
           ]
          print(",".join(column_names))
          print("\n")
        print("\n")


      except:
        print("Error!! there is no work-book")

    return self.__sheet_name




  def extract_data(self, workbook_path, sheet_name):
    try:
      self.clear_data()
      if sheet_name in self.__sheet_name[workbook_path]:
        self.__sheet_data[sheet_name]=None
        df = pd.read_excel(workbook_path, sheet_name=sheet_name,  header=0,engine='openpyxl')
        self.__sheet_data[sheet_name]=df
        print(f"Sucessfully!! Data Extracted from sheet name :{sheet_name}")

      else:
        print("NO Sheet found!!")
    except Exception as e:
      print(f"An error occurred while extracting data: {e}")

  def clear_data(self):
    self.__sheet_data.clear()

  def get_sheet_data(self):
    return self.__sheet_data


  def preview_data(self,rows=5):

    if self.__sheet_data:
      for sheet_name, df in self.__sheet_data.items():
        print(f"Store data of {sheet_name} is ::")
        display(df.head(rows))
        print("\n")

